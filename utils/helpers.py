import asyncio
import sys
from pathlib import Path
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl import Workbook

from .app_config import AppDataConfig, read_config
from .auth import build_ntlm_auth
from .constants import HEADERS, MAIN_URL


async def get_response(
    link: str, config: AppDataConfig | None = None
) -> requests.Response:
    """
    Send an asynchronous GET request to the specified link.

    Args:
        link (str): The URL to send the request to.

    Returns:
        requests.Response: The response object.
    """
    # Run requests in thread pool to make it async-compatible
    loop = asyncio.get_event_loop()
    config = config or read_config()
    auth = build_ntlm_auth(config)
    response = await loop.run_in_executor(
        None, lambda: requests.get(link, headers=HEADERS, auth=auth)
    )
    return response


def get_url_period_equipment_data(link_up: str, date: str, shift: str) -> str:
    """
    Generate a URL for fetching stop data.

    Args:
        link_up (str): The line identifier.
        date (str): The date for the query.
        shift (str): The shift for the query.

    Returns:
        str: The generated URL.
    """
    line_prefix = "PMID-SE-CP-L0" if link_up == "17" else "ID01-SE-CP-L0"
    params = {
        "table": "sp_PeriodEquipmentData",
        "act": "query",
        "eoa": "x",
        "db_Line": f"{line_prefix}{link_up}",
        "db_SegmentDateMin": date,
        "db_ShiftStart": shift,
        "db_ShiftEnd": shift,
    }
    return MAIN_URL + "&".join(f"{key}={value}" for key, value in params.items())


def get_url_period_loss_tree(
    link_up: str, date: str, shift: str = "", functional_location: str = "PACK"
) -> str:
    line_prefix = "PMID-SE-CP-L0" if link_up == "17" else "ID01-SE-CP-L0"
    params = {
        "table": "SPA_NormPeriodLossTree",
        "act": "query",
        "submit1": "Search",
        "db_Line": f"{line_prefix}{link_up}",
        "db_FunctionalLocation": f"{line_prefix}{link_up}-{functional_location}",
        "db_SegmentDateMin": date,
        "db_ShiftStart": shift,
        "db_SegmentDateMax": date,
        "db_ShiftEnd": shift,
        "db_Normalize": 0,
        "db_PeriodTime": 10080,
        "s_PeriodTime": "",
        "db_LongStopDetails": 3,
        "db_ReasonCNT": 30,
        "db_ReasonSort": "stop count",
        "db_Language": "OEM",
        "db_LineFailureAnalysis": "x",
    }

    """http://ots.app.pmi/db.aspx?table=SPA_NormPeriodLossTree&act=query&submit1=Search&db_Line=ID01-SE-CP-L021&db_FunctionalLocation=ID01-SE-CP-L021-MAKE&db_SegmentDateMin=2025-11-08&db_ShiftStart=&db_SegmentDateMax=&db_ShiftEnd=&db_Normalize=0&db_PeriodTime=10080&s_PeriodTime=&db_LongStopDetails=3&db_ReasonCNT=30&db_ReasonSort=stop+count&db_Language=OEM&db_LineFailureAnalysis=x"""

    return MAIN_URL + urlencode(params, doseq=True)


def get_url_norm_period_loss_tree(
    link_up: str, date: str, shift: str, functional_location: str = "PACK"
) -> str:
    """
    Generate a URL for fetching result data.

    Args:
        link_up (str): The line identifier.
        date (str): The date for the query.
        shift (str): The shift for the query.

    Returns:
        str: The generated URL.
    """
    line_prefix = "PMID-SE-CP-L0" if link_up == "17" else "ID01-SE-CP-L0"
    params = {
        "table": "SPA_NormPeriodLossTree",
        "act": "query",
        "db_Normalize": 0,
        "eoa": "x",
        "db_SegmentDateMin": date,
        "db_ShiftStart": shift,
        "db_ShiftEnd": shift,
        "db_Line": f"{line_prefix}{link_up}",
        "db_Language": "OEM",
        "db_FunctionalLocation": f"{line_prefix}{link_up}-{functional_location}",
    }
    """
    table=SPA_NormPeriodLossTree
    &act=query
    &db_Normalize=0
    &eoa=x
    &db_SegmentDateMin=2025-07-10
    &db_ShiftStart=1
    &db_ShiftEnd=1
    &db_Line=ID01-SE-CP-L021
    &db_Language=OEM
    &db_FunctionalLocation=ID01-SE-CP-L021-MAKE

    http://ots.app.pmi/db.aspx?
    table=
    &act=query
    &db_Normalize=0
    &eoa=
    &db_SegmentDateMin=2025-07-10
    &db_ShiftStart=3
    &db_ShiftEnd=3
    &db_Line=PMID-SE-CP-L018
    &db_Language=OEM
    &db_FunctionalLocation=PMID-SE-CP-L018-PACK
    """
    return MAIN_URL + "&".join(f"{key}={value}" for key, value in params.items())


def resource_path(relative_path: str) -> str:
    """
    Get the absolute path to a resource, compatible with PyInstaller.

    Args:
        relative_path (str): The relative path to the resource.

    Returns:
        str: The absolute path to the resource.
    """
    base_path = getattr(sys, "_MEIPASS", str(Path(".").resolve()))
    return str(Path(base_path) / relative_path)


def get_script_folder() -> str:
    """
    Get the absolute path to the script folder, compatible with PyInstaller.

    Returns:
        str: The absolute path to the script folder.
    """
    if getattr(sys, "frozen", False):
        return str(Path(sys.executable).parent)
    return str(Path(sys.modules["__main__"].__file__).resolve().parent)


def create_excel_file() -> None:
    """
    Create an Excel file with predefined sheets if it doesn't already exist.
    """
    file_path = Path(get_script_folder()) / "DB.xlsx"
    sheets_list = ["Data", "Username", "Link", "DailyTarget"]

    if not file_path.exists():
        wb = Workbook()
        wb.active.title = "Data"
        for sheet_name in sheets_list[1:]:
            wb.create_sheet(title=sheet_name)

        # Add default data to the "DailyTarget" sheet
        ws = wb["DailyTarget"]
        ws.append(["", "TARGET"])
        for index in ["STOP", "MTBF", "PR", "NATR", "PDT", "UPDT"]:
            ws.append([index])

        wb.save(file_path)


def get_excel_filename() -> str:
    """
    Get the path to the Excel file, creating it if it doesn't exist.

    Returns:
        str: The path to the Excel file.
    """
    file_path = Path(get_script_folder()) / "DB.xlsx"
    if not file_path.exists():
        create_excel_file()
    return str(file_path)


def get_data_from_excel(sheet_index: int):
    """
    Retrieve data from a specific sheet in the Excel file.

    Args:
        sheet_index (int): The index of the sheet to read.

    Returns:
        list: A list of values from the first column of the sheet.
    """
    file_path = get_excel_filename()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[sheet_index]

    data = [row[0].value for row in sheet.iter_rows() if row[0].value]
    wb.close()
    return data
